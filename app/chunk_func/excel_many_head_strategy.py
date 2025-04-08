from typing import List, Dict, Any, Tuple, Optional
import pandas as pd
import json
import logging
import os
import numpy as np
import openpyxl
from .base import BaseChunkStrategy

# 设置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelManyHeadStrategy(BaseChunkStrategy):
    """Excel多级表头识别切块策略，支持2行和3行表头"""
    
    def __init__(self):
        super().__init__()
    
    def _detect_header_rows(self, df_sample: pd.DataFrame, max_rows: int = 20) -> Tuple[int, int, int]:
        """
        检测表头行范围，区分二级表头和三级表头
        
        Args:
            df_sample: 数据帧样本
            max_rows: 最多检查的行数
            
        Returns:
            (表头开始行, 表头结束行, 表头级别(2或3))
        """
        num_rows = min(max_rows, df_sample.shape[0])
        if num_rows <= 3:  # 至少需要3行来判断表头
            return 0, 2, 3
        
        # 计算每行的特征
        row_features = []
        for i in range(num_rows):
            row = df_sample.iloc[i]
            # 计算非空率
            non_empty_count = sum(1 for x in row if not pd.isna(x))
            non_empty_ratio = non_empty_count / len(row) if len(row) > 0 else 0
            
            # 计算数字比例
            numeric_count = sum(isinstance(x, (int, float, np.integer, np.floating)) 
                              for x in row if not pd.isna(x))
            numeric_ratio = numeric_count / non_empty_count if non_empty_count > 0 else 0
            
            row_features.append({
                'non_empty_ratio': non_empty_ratio,
                'numeric_ratio': numeric_ratio
            })
        
        # 首先尝试检测三级表头
        for i in range(num_rows - 2):
            # 表头特征：非空率较高（>0.3），数字比例较低（<0.3）
            if all(
                row_features[j]['non_empty_ratio'] > 0.3 and 
                row_features[j]['numeric_ratio'] < 0.3
                for j in range(i, i + 3)
            ):
                # 检查下一行是否为数据行（数字比例较高）
                if (i + 3 < num_rows and 
                    row_features[i + 3]['numeric_ratio'] > 0.5):
                    logger.info(f"检测到三级表头范围: 第{i}行到第{i+2}行")
                    return i, i + 2, 3
        
        # 如果没有检测到三级表头，尝试检测二级表头
        for i in range(num_rows - 1):
            if (row_features[i]['numeric_ratio'] < 0.3 and
                row_features[i]['non_empty_ratio'] > 0.3 and 
                row_features[i+1]['non_empty_ratio'] > 0.3 and 
                row_features[i+1]['numeric_ratio'] < 0.3):
                # 检查下一行是否为数据行
                if (i + 2 < num_rows and 
                    row_features[i + 2]['numeric_ratio'] > 0.5):
                    logger.info(f"检测到二级表头范围: 第{i}行到第{i+1}行")
                    return i, i + 1, 2
        
        # 如果没有找到明显的特征，默认使用前三行作为三级表头
        logger.info("未检测到明显的表头特征，默认使用前三行作为三级表头")
        return 0, 2, 3
    
    def _get_merged_cell_value(self, sheet, row, col) -> str:
        """获取合并单元格的值"""
        cell = sheet.cell(row=row, column=col)
        if cell.value is not None:
            return str(cell.value).strip()
            
        # 检查是否在合并单元格内
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
            if min_row <= row <= max_row and min_col <= col <= max_col:
                return str(sheet.cell(row=min_row, column=min_col).value or "").strip()
        return ""
    
    def _get_merged_range(self, sheet, row, col) -> Optional[Tuple[int, int, int, int]]:
        """获取单元格所在的合并区域范围"""
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
            if min_row <= row <= max_row and min_col <= col <= max_col:
                return (min_row, min_col, max_row, max_col)
        return None
    
    def _analyze_two_level_header(self, sheet, col_idx, header_start) -> str:
        """
        分析两行表头的单列的表头结构
        
        处理两种情况：
        1. 两行都合并了，这两行都是一个值
        2. 第一行一个值；第二行一个值
        """
        row1 = header_start
        row2 = header_start + 1
        
        # 获取两行的合并区域信息
        range1 = self._get_merged_range(sheet, row1, col_idx)
        range2 = self._get_merged_range(sheet, row2, col_idx)
        
        # 获取两行的值
        value1 = self._get_merged_cell_value(sheet, row1, col_idx)
        value2 = self._get_merged_cell_value(sheet, row2, col_idx)
        
        # 情况1：两行都合并了
        if (range1 and range1 == range2 and 
            range1[0] == row1 and range1[2] == row2):
            return value1
        
        # 情况2：两行都是独立的值
        if ((not range1 or range1[0] == range1[2]) and 
            (not range2 or range2[0] == range2[2])):
            # 如果两个值都存在，使用连字符连接
            if value1 and value2:
                return f"{value1}-{value2}"
            # 如果只有一个值存在，直接返回该值
            elif value1:
                return value1
            elif value2:
                return value2
            else:
                return f"列{col_idx}"
        
        # 其他情况：尽可能收集所有不重复的值
        values = []
        if value1 and value1 not in values:
            values.append(value1)
        if value2 and value2 not in values:
            values.append(value2)
        
        return '-'.join(values) if values else f"列{col_idx}"
    
    def _analyze_three_level_header(self, sheet, col_idx, header_start) -> str:
        """
        分析三行表头的单列的表头结构
        
        处理四种情况：
        1. 整三行都合并为一个单元格
        2. 三行分别是独立的值
        3. 第一行独立，二三行合并
        4. 一二行合并，第三行独立
        """
        row1 = header_start
        row2 = header_start + 1
        row3 = header_start + 2
        
        # 获取三行的合并区域信息
        range1 = self._get_merged_range(sheet, row1, col_idx)
        range2 = self._get_merged_range(sheet, row2, col_idx)
        range3 = self._get_merged_range(sheet, row3, col_idx)
        
        # 获取三行的值
        value1 = self._get_merged_cell_value(sheet, row1, col_idx)
        value2 = self._get_merged_cell_value(sheet, row2, col_idx)
        value3 = self._get_merged_cell_value(sheet, row3, col_idx)
        
        # 情况1：整三行都合并了
        if (range1 and range1 == range2 == range3 and 
            range1[0] == row1 and range1[2] == row3):
            return value1
        
        # 情况2：三行都是独立的值
        if ((not range1 or range1[0] == range1[2]) and 
            (not range2 or range2[0] == range2[2]) and 
            (not range3 or range3[0] == range3[2])):
            parts = [v for v in [value1, value2, value3] if v]
            return '-'.join(parts) if parts else f"列{col_idx}"
        
        # 情况3：第一行独立，二三行合并
        if ((not range1 or range1[0] == range1[2]) and 
            range2 and range2 == range3 and 
            range2[0] == row2 and range2[2] == row3):
            parts = [value1, value2] if value1 and value2 else [value1 or value2]
            return '-'.join(parts) if parts else f"列{col_idx}"
        
        # 情况4：一二行合并，第三行独立
        if (range1 and range1 == range2 and 
            range1[0] == row1 and range1[2] == row2 and 
            (not range3 or range3[0] == range3[2])):
            parts = [value1, value3] if value1 and value3 else [value1 or value3]
            return '-'.join(parts) if parts else f"列{col_idx}"
        
        # 其他情况：尽可能收集所有不重复的值
        values = []
        if value1 and value1 not in values:
            values.append(value1)
        if value2 and value2 not in values:
            values.append(value2)
        if value3 and value3 not in values:
            values.append(value3)
        
        return '-'.join(values) if values else f"列{col_idx}"
    
    def chunk_with_meta(self, file_path: str, chunk_size: int, overlap: int) -> List[Dict[str, Any]]:
        """
        将Excel文件按行转换为字典列表，支持2行和3行表头
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
            
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in ['.xlsx', '.xls']:
                raise ValueError(f"不支持的文件类型: {file_ext}")
            
            # 读取前20行用于表头检测
            df_sample = pd.read_excel(file_path, nrows=20)
            header_start, header_end, header_level = self._detect_header_rows(df_sample)
            
            # 使用openpyxl读取Excel文件以支持合并单元格分析
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # 分析每列的表头
            column_names = []
            for col_idx in range(1, sheet.max_column + 1):
                if header_level == 2:
                    header_name = self._analyze_two_level_header(sheet, col_idx, header_start + 1)
                else:  # header_level == 3
                    header_name = self._analyze_three_level_header(sheet, col_idx, header_start + 1)
                column_names.append(header_name)
                logger.info(f"列 {col_idx} 的表头: {header_name}")
            
            # 读取数据行
            result_chunks = []
            for row_idx in range(header_end + 2, sheet.max_row + 1):
                row_dict = {}
                has_data = False
                
                for col_idx in range(1, len(column_names) + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    column_name = column_names[col_idx - 1]
                    
                    if cell.value is not None:
                        has_data = True
                        # 处理数值类型
                        if isinstance(cell.value, (int, float, np.integer, np.floating)):
                            row_dict[column_name] = float(cell.value)
                        else:
                            row_dict[column_name] = str(cell.value).strip()
                    else:
                        row_dict[column_name] = ""
                
                if has_data:
                    result_chunks.append({
                        "content": json.dumps(row_dict, ensure_ascii=False, indent=2),
                        "meta": {
                            "row_index": row_idx,
                            "columns": column_names,
                            "file_name": os.path.basename(file_path),
                            "sheet_name": sheet.title,
                            "header_rows": f"{header_start + 1}-{header_end + 1}",
                            "header_levels": header_level
                        }
                    })
            
            return result_chunks
            
        except Exception as e:
            logger.error(f"Excel多级表头切块过程中发生错误: {str(e)}")
            raise
    
    def get_metadata(self) -> Dict[str, Any]:
        return {
            "name": "excel_many_head",
            "display_name": "Excel2、3级表头识别切块策略",
            "description": "支持2行和3行表头的Excel文件，自动识别表头位置和层级",
            "supported_types": [".xlsx", ".xls"]
        }