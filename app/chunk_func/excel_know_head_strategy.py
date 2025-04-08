from typing import List, Dict, Any, Tuple
import pandas as pd
import json
import logging
import os
import numpy as np
from openpyxl import load_workbook
from .base import BaseChunkStrategy

# 设置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelKnowHeadStrategy(BaseChunkStrategy):
    """Excel智能表头识别切块策略，支持多级表头识别与处理"""
    
    def __init__(self):
        super().__init__()
    
    def _detect_header_row(self, df_sample: pd.DataFrame) -> int:
        """简单表头行检测（用于CSV文件）"""
        # 检查前20行
        num_rows = min(20, df_sample.shape[0])
        if num_rows <= 1:
            return 0
            
        # 计算每行的数字比例
        for i in range(num_rows - 1):
            row1 = df_sample.iloc[i]
            row2 = df_sample.iloc[i+1]
            
            # 计算当前行和下一行的数字比例
            r1_nums = sum(isinstance(x, (int, float, np.integer, np.floating)) for x in row1 if not pd.isna(x))
            r1_total = sum(1 for x in row1 if not pd.isna(x))
            r1_ratio = r1_nums / r1_total if r1_total > 0 else 0
            
            r2_nums = sum(isinstance(x, (int, float, np.integer, np.floating)) for x in row2 if not pd.isna(x))
            r2_total = sum(1 for x in row2 if not pd.isna(x))
            r2_ratio = r2_nums / r2_total if r2_total > 0 else 0
            
            # 表头后通常数字比例增加
            if r1_ratio < 0.3 and r2_ratio > 0.5:
                return i
        
        return 0
    
    def _process_excel_file(self, file_path: str) -> Tuple[pd.DataFrame, Dict]:
        """处理Excel文件，识别表头并返回DataFrame"""
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # 检查前20行，找出表头行
        header_rows = []
        for row_idx in range(1, min(21, ws.max_row + 1)):
            # 检查行属性
            row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
            non_empty = sum(1 for x in row_values if x)
            if non_empty / len(row_values) < 0.5:
                continue
                
            numeric_count = sum(1 for x in row_values if isinstance(x, (int, float)))
            if numeric_count / non_empty > 0.5:
                continue
                
            header_rows.append(row_idx - 1)  # 转为0-based索引
        
        # 找出连续的表头行
        if not header_rows:
            # 没找到合适的表头行，使用第一行
            df = pd.read_excel(file_path, header=0)
            return df, {"type": "single", "header_row": 0}
        
        # 确定表头范围
        header_start = header_rows[0]
        header_end = header_start
        
        for i in range(1, len(header_rows)):
            if header_rows[i] == header_rows[i-1] + 1:
                header_end = header_rows[i]
            else:
                break
        
        if header_start == header_end:
            # 单行表头
            df = pd.read_excel(file_path, header=header_start)
            return df, {"type": "single", "header_row": header_start}
        
        # 处理多行表头
        # 获取表头数据和合并单元格信息
        merged_cells = ws.merged_cells.ranges
        merged_info = {}
        parent_headers = {}
        
        # 处理合并单元格
        for merged_range in merged_cells:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            
            # 只关注表头区域的合并单元格
            if not (header_start + 1 <= min_row <= header_end + 1 and header_start + 1 <= max_row <= header_end + 1):
                continue
                
            # 获取合并单元格值
            value = ws.cell(row=min_row, column=min_col).value
            if value is None:
                value = ""
            
            # 记录合并信息
            if min_row == max_row:  # 列合并
                # 如果是第一行，记录为父级表头
                if min_row == header_start + 1:
                    for col in range(min_col, max_col + 1):
                        parent_headers[col] = str(value).strip()
            else:  # 行合并
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        if r != min_row or c != min_col:
                            merged_info[(r, c)] = (min_row, min_col, value)
        
        # 构建表头
        header_names = []
        
        for col in range(1, ws.max_column + 1):
            parts = []
            
            # 添加父级表头
            if col in parent_headers and parent_headers[col]:
                parts.append(parent_headers[col])
            
            # 处理每行表头
            for row in range(header_start + 1, header_end + 2):
                # 跳过合并单元格
                if (row, col) in merged_info:
                    _, _, value = merged_info[(row, col)]
                    if value and value not in parts:
                        parts.append(str(value).strip())
                    continue
                
                # 获取单元格值
                value = ws.cell(row=row, column=col).value
                if value is not None and str(value).strip():
                    if str(value).strip() not in parts:
                        parts.append(str(value).strip())
            
            # 创建表头名称
            header_name = "-".join(parts) if parts else f"Column_{col-1}"
            header_names.append(header_name)
        
        # 读取数据
        df = pd.read_excel(file_path, header=header_end)
        
        # 重命名列
        if len(header_names) == len(df.columns):
            df.columns = header_names
        
        return df, {
            "type": "multi_level", 
            "header_start": header_start,
            "header_end": header_end,
            "headers": header_names
        }
    
    def chunk_with_meta(self, file_path: str, chunk_size: int, overlap: int) -> List[Dict[str, Any]]:
        """将Excel文件按行转换为字典列表，支持多级表头"""
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
            
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in ['.xlsx', '.xls', '.csv']:
                raise ValueError(f"不支持的文件类型: {file_ext}")
            
            # 处理文件
            if file_ext == '.csv':
                # CSV文件使用简单表头识别
                df_sample = pd.read_csv(file_path, nrows=20)
                header_row = self._detect_header_row(df_sample)
                df = pd.read_csv(file_path, header=header_row)
                header_info = {"type": "simple", "header_row": header_row}
            else:
                # Excel文件使用多级表头识别
                df, header_info = self._process_excel_file(file_path)
            
            if df.empty:
                return []
            
            # 转换为字典列表
            result_chunks = []
            for index, row in df.iterrows():
                # 处理数据
                row_dict = {}
                for k, v in row.to_dict().items():
                    if pd.isna(v):
                        row_dict[k] = ""
                    elif isinstance(v, (np.integer, np.floating)):
                        row_dict[k] = float(v) if isinstance(v, np.floating) else int(v)
                    else:
                        row_dict[k] = v
                
                # 添加元数据
                meta = {
                    "row_index": int(index) if isinstance(index, (np.integer, np.int64)) else index,
                    "file_name": os.path.basename(file_path),
                    "header_info": header_info
                }
                
                result_chunks.append({
                    "content": json.dumps(row_dict, ensure_ascii=False, indent=2),
                    "meta": meta
                })
            
            return result_chunks
            
        except Exception as e:
            logger.error(f"Excel切块过程中发生错误: {str(e)}")
            raise
    
    def get_metadata(self) -> Dict[str, Any]:
        return {
            "name": "excel_know_head",
            "display_name": "Excel智能表头识别切块策略",
            "description": "自动识别Excel表头位置，支持多级表头，将每行数据转换为键值对形式的切片",
            "supported_types": [".xlsx", ".xls", ".csv"]
        }