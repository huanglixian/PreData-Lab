from typing import List, Dict, Any, Tuple
import pandas as pd
import json
import logging
import os
import numpy as np
from openpyxl import load_workbook
from .base import BaseChunkStrategy

# 设置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class Excel3HeadStrategy(BaseChunkStrategy):
    """Excel三级表头识别切块策略，针对特定的三级表头结构进行优化识别"""
    
    def __init__(self):
        super().__init__()
    
    def _detect_header_row(self, df_sample: pd.DataFrame) -> int:
        """简单表头行检测（用于CSV文件）"""
        # 检查前20行
        num_rows = min(20, df_sample.shape[0])
        if num_rows <= 1:
            return 0
            
        # 计算每行的数字比例
        for i in range(num_rows - 1):
            row1 = df_sample.iloc[i]
            row2 = df_sample.iloc[i+1]
            
            # 计算当前行和下一行的数字比例
            r1_nums = sum(isinstance(x, (int, float, np.integer, np.floating)) for x in row1 if not pd.isna(x))
            r1_total = sum(1 for x in row1 if not pd.isna(x))
            r1_ratio = r1_nums / r1_total if r1_total > 0 else 0
            
            r2_nums = sum(isinstance(x, (int, float, np.integer, np.floating)) for x in row2 if not pd.isna(x))
            r2_total = sum(1 for x in row2 if not pd.isna(x))
            r2_ratio = r2_nums / r2_total if r2_total > 0 else 0
            
            # 表头后通常数字比例增加
            if r1_ratio < 0.3 and r2_ratio > 0.5:
                return i
        
        return 0
    
    def _process_excel_file(self, file_path: str) -> Tuple[pd.DataFrame, Dict]:
        """处理Excel文件，识别三级表头并返回DataFrame"""
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # 检查前20行，找出表头行
        header_rows = []
        for row_idx in range(1, min(21, ws.max_row + 1)):
            # 检查行属性
            row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
            non_empty = sum(1 for x in row_values if x)
            if non_empty / len(row_values) < 0.5:
                continue
                
            numeric_count = sum(1 for x in row_values if isinstance(x, (int, float)))
            if numeric_count / non_empty > 0.5:
                continue
                
            header_rows.append(row_idx - 1)  # 转为0-based索引
        
        # 找出连续的表头行
        if not header_rows:
            # 没找到合适的表头行，使用第一行
            df = pd.read_excel(file_path, header=0)
            return df, {"type": "single", "header_row": 0}
        
        # 确定表头范围
        header_start = header_rows[0]
        header_end = header_start
        
        for i in range(1, len(header_rows)):
            if header_rows[i] == header_rows[i-1] + 1:
                header_end = header_rows[i]
            else:
                break
        
        if header_start == header_end:
            # 单行表头
            df = pd.read_excel(file_path, header=header_start)
            return df, {"type": "single", "header_row": header_start}
        
        # 判断是否是三级表头
        if header_end - header_start == 2:
            logger.info("检测到三级表头结构")
        else:
            logger.info(f"检测到{header_end - header_start + 1}级表头结构")
        
        # 处理多行表头
        merged_cells = ws.merged_cells.ranges
        merged_info = {}  # 记录被合并的单元格位置
        
        # 首先分析所有的合并单元格区域
        merged_regions = []
        for merged_range in merged_cells:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            
            # 只关注表头区域的合并单元格
            if not (header_start + 1 <= min_row <= header_end + 1 and min_row <= max_row <= header_end + 1):
                continue
            
            # 获取合并单元格值
            value = ws.cell(row=min_row, column=min_col).value
            if value is None:
                value = ""
            else:
                value = str(value).strip()
            
            # 记录合并区域信息
            merged_regions.append({
                'min_row': min_row,
                'max_row': max_row,
                'min_col': min_col,
                'max_col': max_col,
                'value': value,
                'level': min_row - (header_start + 1),  # 级别
                'span_type': 'horizontal' if min_row == max_row else 'vertical' if min_col == max_col else 'area'
            })
            
            # 标记被合并的单元格
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if r != min_row or c != min_col:
                        merged_info[(r, c)] = (min_row, min_col, value)
        
        # 对合并区域进行排序处理
        # 先按级别排序，同级别的水平合并优先于垂直合并
        merged_regions.sort(key=lambda x: (x['level'], 0 if x['span_type'] == 'horizontal' else 1))
        
        # 识别各列所属的分组
        col_groups = self._identify_column_groups(merged_regions, ws.max_column)
        
        # 生成三级表头结构
        level_headers = [{} for _ in range(3)]  # 三级表头
        
        # 1. 从合并单元格中提取信息
        for region in merged_regions:
            level = region['level']
            if level >= 3:  # 超过三级表头的忽略
                continue
                
            value = region['value']
            if not value:  # 跳过空值
                continue
                
            # 根据合并单元格类型处理
            if region['span_type'] == 'horizontal':  # 水平合并（跨列）
                # 这可能是一个分组表头
                for col in range(region['min_col'], region['max_col'] + 1):
                    level_headers[level][col] = value
            else:  # 垂直合并或区域合并
                # 记录此单元格的值
                level_headers[level][region['min_col']] = value
        
        # 2. 处理未合并的单元格
        for level in range(3):
            row = header_start + 1 + level
            if row > header_end + 1:  # 超出表头范围
                continue
                
            for col in range(1, ws.max_column + 1):
                # 跳过已处理的合并单元格
                if (row, col) in merged_info:
                    continue
                    
                value = ws.cell(row=row, column=col).value
                if value is not None:
                    value = str(value).strip()
                    if value:
                        level_headers[level][col] = value
        
        # 3. 特殊处理表格特定区域
        # 这里进行针对性的优化，根据用户描述的表格结构
        
        # 处理列1-3（基本信息列）
        for col in range(1, 4):
            # 确保这些列有合适的值
            if col == 1 and col not in level_headers[0]:
                level_headers[0][col] = "序号"
            elif col == 2 and col not in level_headers[0]:
                level_headers[0][col] = "编制依据"
            elif col == 3 and col not in level_headers[0]:
                level_headers[0][col] = "项目名称规范"
        
        # 处理列4-5（量价列）
        if 4 in level_headers[0] and 5 in level_headers[0]:
            # 检查是否有共同的第一级
            if level_headers[0][4] == level_headers[0][5]:
                # 有共同的第一级（比如"量价"）
                pass
            else:
                # 没有共同的第一级，添加"量价"作为共同第一级
                common_prefix = "量价"
                level_headers[0][4] = level_headers[0][5] = common_prefix
        
        # 为列4-5添加第二级名称
        if 4 not in level_headers[1] or not level_headers[1][4]:
            level_headers[1][4] = "单位"
        if 5 not in level_headers[1] or not level_headers[1][5]:
            level_headers[1][5] = "数量"
        
        # 处理列6-11（金额列）
        # 第一级：金额
        for col in range(6, 12):
            if col not in level_headers[0] or not level_headers[0][col]:
                level_headers[0][col] = "金额"
        
        # 第二级：单价（列6-8）和合价（列9-11）
        for col in range(6, 9):
            if col not in level_headers[1] or not level_headers[1][col]:
                level_headers[1][col] = "单价"
        
        for col in range(9, 12):
            if col not in level_headers[1] or not level_headers[1][col]:
                level_headers[1][col] = "合价"
        
        # 第三级：合计、其中工资、其中机械
        third_level_names = ["合     计", "其中工资", "其中机械"]
        for i, col in enumerate(range(6, 9)):
            if i < len(third_level_names):
                if col not in level_headers[2] or not level_headers[2][col]:
                    level_headers[2][col] = third_level_names[i]
        
        for i, col in enumerate(range(9, 12)):
            if i < len(third_level_names):
                if col not in level_headers[2] or not level_headers[2][col]:
                    level_headers[2][col] = third_level_names[i]
        
        # 构建最终表头
        header_names = []
        for col in range(1, ws.max_column + 1):
            parts = []
            
            # 收集各级表头
            for level in range(3):
                if col in level_headers[level] and level_headers[level][col]:
                    parts.append(level_headers[level][col])
            
            # 创建组合表头名称
            if parts:
                header_name = "-".join(parts)
            else:
                header_name = f"Column_{col-1}"
            
            # 检查重复并添加
            base_name = header_name
            count = 1
            while header_name in header_names:
                header_name = f"{base_name}.{count}"
                count += 1
            
            header_names.append(header_name)
        
        # 读取数据
        df = pd.read_excel(file_path, header=header_end)
        
        # 重命名列
        if len(header_names) == len(df.columns):
            df.columns = header_names
        else:
            logger.warning(f"列名数量({len(header_names)})与DataFrame列数({len(df.columns)})不匹配，使用默认列名")
        
        return df, {
            "type": "multi_level", 
            "header_start": header_start,
            "header_end": header_end,
            "headers": header_names
        }
    
    def _identify_column_groups(self, merged_regions, max_columns):
        """基于合并单元格识别列分组"""
        groups = []
        
        # 提取水平合并的单元格（可能是列分组）
        h_merged = [r for r in merged_regions if r['span_type'] == 'horizontal']
        
        # 按列范围排序
        h_merged.sort(key=lambda x: (x['min_col'], -x['max_col']))
        
        # 识别列分组
        for region in h_merged:
            min_col = region['min_col']
            max_col = region['max_col']
            level = region['level']
            value = region['value']
            
            # 只考虑跨越多个列的合并单元格
            if max_col > min_col:
                # 检查是否是已知的特殊分组
                if min_col <= 5 and max_col >= 4:
                    # 可能是"量价"分组
                    groups.append({
                        'start': min_col,
                        'end': max_col,
                        'level': level,
                        'name': value if value else "量价"
                    })
                elif min_col >= 6 and max_col <= 11:
                    # 可能是"金额"下的分组
                    if level == 0:
                        groups.append({
                            'start': min_col,
                            'end': max_col,
                            'level': level,
                            'name': value if value else "金额"
                        })
                    elif level == 1:
                        # 区分"单价"和"合价"分组
                        if min_col <= 8:
                            groups.append({
                                'start': min_col,
                                'end': min(max_col, 8),
                                'level': level,
                                'name': value if value else "单价"
                            })
                        if max_col >= 9:
                            groups.append({
                                'start': max(min_col, 9),
                                'end': max_col,
                                'level': level,
                                'name': value if value else "合价"
                            })
                else:
                    # 其他分组
                    groups.append({
                        'start': min_col,
                        'end': max_col,
                        'level': level,
                        'name': value
                    })
        
        return groups
    
    def chunk_with_meta(self, file_path: str, chunk_size: int, overlap: int) -> List[Dict[str, Any]]:
        """将Excel文件按行转换为字典列表，支持三级表头"""
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
            
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in ['.xlsx', '.xls', '.csv']:
                raise ValueError(f"不支持的文件类型: {file_ext}")
            
            # 处理文件
            if file_ext == '.csv':
                # CSV文件使用简单表头识别
                df_sample = pd.read_csv(file_path, nrows=20)
                header_row = self._detect_header_row(df_sample)
                df = pd.read_csv(file_path, header=header_row)
                header_info = {"type": "simple", "header_row": header_row}
            else:
                # Excel文件使用三级表头识别
                df, header_info = self._process_excel_file(file_path)
            
            if df.empty:
                return []
            
            # 转换为字典列表
            result_chunks = []
            for index, row in df.iterrows():
                # 处理数据
                row_dict = {}
                for k, v in row.to_dict().items():
                    if pd.isna(v):
                        row_dict[k] = ""
                    elif isinstance(v, (np.integer, np.floating)):
                        row_dict[k] = float(v) if isinstance(v, np.floating) else int(v)
                    else:
                        row_dict[k] = v
                
                # 添加元数据
                meta = {
                    "row_index": int(index) if isinstance(index, (np.integer, np.int64)) else index,
                    "file_name": os.path.basename(file_path),
                    "header_info": header_info
                }
                
                result_chunks.append({
                    "content": json.dumps(row_dict, ensure_ascii=False, indent=2),
                    "meta": meta
                })
            
            return result_chunks
            
        except Exception as e:
            logger.error(f"Excel切块过程中发生错误: {str(e)}")
            raise
    
    def get_metadata(self) -> Dict[str, Any]:
        return {
            "name": "excel_3head",
            "display_name": "Excel三级表头切块策略",
            "description": "专门针对三级表头结构设计，能够精确识别复杂表头关系，不依赖映射配置",
            "supported_types": [".xlsx", ".xls", ".csv"]
        } 